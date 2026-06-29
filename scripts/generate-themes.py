#!/usr/bin/env python3
"""
Generate broad themes from clustered codes — thesis-ready output.

Reads from: "Ojectives matrix Final.xlsx" (the cleaned clustered export)

3-turn LLM approach (per objective×category slice, then consolidate):
  Turn 1: For each objective×category combo (~15-37 rows), propose 2-3 themes
  Turn 2: Consolidate the per-objective themes into 4-6 final themes
  Turn 3: Assign every cluster row to a final theme + pick representative excerpts

Produces:
  - thematic-analysis.xlsx (4 sheets: P+S Fac, P+S Bar, Doc Fac, Doc Bar)
  - themed-analysis-raw.json (full LLM output for inspection)

Total LLM calls: up to 6 (Turn 1) + 1 (Turn 2) + up to 3 (Turn 3) = ~10 per section × 4 = ~40
"""
import json, re
from pathlib import Path
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import anthropic

ROOT = Path(__file__).parent.parent
INPUT_XLSX = ROOT / 'Ojectives matrix Final.xlsx'

# ── Load API key ──
env_vars = {}
for line in (ROOT / '.env.local').read_text().splitlines():
    line = line.strip()
    if '=' in line and not line.startswith('#'):
        k, v = line.split('=', 1)
        env_vars[k.strip()] = v.strip()

api_key = env_vars['ANTHROPIC_API_KEY']

# ══════════════════════════════════════════════════════════════
# STEP 1: Parse the input Excel into structured data
# ══════════════════════════════════════════════════════════════

print('Reading input Excel...')
wb_in = load_workbook(INPUT_XLSX, read_only=True)
ws = wb_in['Clustered Themes']

# Parse rows into structured sections
# Each row (after section headers) has: [Theme/ClusterName, Coverage, Participants, Findings]
current_type = None   # 'doctor', 'patient', 'survivor'
current_obj = None    # 'objective_1', 'objective_2', 'objective_3'
current_cat = None    # 'facilitator', 'barrier'

# Storage: { (group, obj, cat) : [ {name, coverage, participants, findings} ] }
# group = 'patient_survivor' or 'doctor'
parsed_sections = defaultdict(list)

OBJ_MAP = {
    'Objective 1': 'objective_1',
    'Objective 2': 'objective_2',
    'Objective 3': 'objective_3',
}

for row in ws.iter_rows(values_only=True):
    if not row or len(row) < 4:
        continue
    a = str(row[0]).strip() if row[0] else ''

    # Detect type header: "Doctors (6 interviews)"
    if 'interviews)' in a:
        if 'Doctor' in a:
            current_type = 'doctor'
        elif 'Patient' in a:
            current_type = 'patient'
        elif 'Survivor' in a:
            current_type = 'survivor'
        continue

    # Detect objective+category header: "Objective 1 — Early Detection  ·  FACILITATORS"
    if 'Objective' in a and ('FACILITATOR' in a or 'BARRIER' in a):
        for prefix, obj_key in OBJ_MAP.items():
            if prefix in a:
                current_obj = obj_key
                break
        current_cat = 'facilitator' if 'FACILITATOR' in a else 'barrier'
        continue

    # Skip header rows and empty rows
    if not a or a in ('Theme', 'Findings', 'Cluster / Theme'):
        continue

    if current_type is None or current_obj is None or current_cat is None:
        continue

    # Data row
    coverage = str(row[1]).strip() if row[1] else ''
    participants = str(row[2]).strip() if row[2] else ''
    findings = str(row[3]).strip() if row[3] else ''

    # Determine group
    group = 'patient_survivor' if current_type in ('patient', 'survivor') else 'doctor'

    parsed_sections[(group, current_obj, current_cat)].append({
        'name': a,
        'coverage': coverage,
        'participants': participants,
        'findings': findings,
        'source_type': current_type,
    })

wb_in.close()

# Print summary
total_rows = 0
for key in sorted(parsed_sections.keys()):
    count = len(parsed_sections[key])
    total_rows += count
    print(f'  {key}: {count} cluster rows')
print(f'  Total: {total_rows} rows')

# ══════════════════════════════════════════════════════════════
# STEP 2: LLM calls
# ══════════════════════════════════════════════════════════════

client = anthropic.Anthropic(api_key=api_key)
MODEL = 'claude-haiku-4-5-20251001'

OBJ_LABELS = {
    'objective_1': 'Objective 1 — Early Detection',
    'objective_2': 'Objective 2 — Diagnosis & Treatment',
    'objective_3': 'Objective 3 — Continuity & Follow-Up',
}
GROUP_LABELS = {
    'patient_survivor': 'Patients & Survivors',
    'doctor': 'Doctors (Oncologists)',
}
CAT_LABELS = {
    'facilitator': 'Facilitators',
    'barrier': 'Barriers',
}

STUDY_CONTEXT = """STUDY BACKGROUND:
This is a qualitative cross-sectional study examining the breast cancer patient care pathway in Hyderabad, India. Semi-structured interviews were conducted with three participant groups: patients (women currently undergoing treatment), survivors (women who completed treatment), and oncologists (doctors).

The study examines facilitators (enabling factors) and barriers (hindering factors) across three stages of the care pathway:
- Objective 1 — Early Detection: awareness, screening practices, symptom recognition, health-seeking behaviour
- Objective 2 — Diagnosis & Treatment: diagnostic workup, treatment initiation, treatment delivery, decision-making
- Objective 3 — Continuity & Follow-Up: treatment adherence, follow-up care, psychosocial support, survivorship, quality of life

The data has already been coded and clustered into sub-themes through a systematic process. Your role now is to identify broader overarching themes that group these sub-themes."""


def llm_call(system, user, max_tokens=4096, retries=2):
    """Make a single LLM call with JSON extraction and retry on parse failure."""
    for attempt in range(retries + 1):
        try:
            response = client.messages.create(
                model=MODEL,
                max_tokens=max_tokens,
                system=system,
                messages=[
                    {"role": "user", "content": user},
                    {"role": "assistant", "content": "{"},
                ],
            )
            raw = '{' + response.content[0].text
            # Strip trailing text after the last closing brace (LLM sometimes appends commentary)
            last_brace = raw.rfind('}')
            if last_brace != -1:
                raw = raw[:last_brace + 1]
            data = json.loads(raw)
            tokens_in = response.usage.input_tokens
            tokens_out = response.usage.output_tokens
            return data, tokens_in, tokens_out
        except (json.JSONDecodeError, KeyError, IndexError) as e:
            if attempt < retries:
                print(f'  ⚠ JSON parse error (attempt {attempt+1}), retrying... ({e})')
            else:
                print(f'  ✗ JSON parse failed after {retries+1} attempts. Raw response:')
                print(f'    {raw[:500]}')
                raise


# ────────────────────────────────────────
# Turn 1: Per-objective theme proposals
# ────────────────────────────────────────

TURN1_SYSTEM = f"""{STUDY_CONTEXT}

TASK: You will receive a list of sub-themes (codes/clusters) from ONE specific objective and category (e.g. "Objective 1 Facilitators"). Each sub-theme has a name, coverage (how many interviews mentioned it), and the individual findings within it.

Your job: propose 2-3 BROAD themes that meaningfully group these sub-themes together.

STRICT RULES:
1. GROUNDED ONLY: Every theme you propose must directly emerge from the sub-themes provided. Do NOT introduce concepts, theories, or ideas not present in the data.
2. COVERAGE MATTERS: Themes should prioritise sub-themes with higher coverage (mentioned by more participants). Single-participant sub-themes can be grouped but should not drive theme naming.
3. MUTUAL EXCLUSIVITY: Each sub-theme should clearly belong to only one theme. If you cannot decide, place it in the more specific theme.
4. NAMING: Theme names should be 5-12 words, academic but clear. They should read well as section headings in a thesis chapter.
5. EXACTLY 2-3 themes. Not more, not fewer.
6. PATIENT + SURVIVOR OVERLAP: When the participant group is "Patients & Survivors", you will see sub-themes from both patients and survivors. Many will be semantically similar (e.g. "Lack of knowledge about breast cancer" from patients and "Lack of breast cancer awareness" from survivors). Treat them as ONE cohort — do not create separate themes for patients vs survivors.

OUTPUT FORMAT — return ONLY this JSON:
{{
  "themes": [
    {{
      "name": "Proposed theme name",
      "rationale": "1-2 sentences explaining what sub-themes this covers and why they belong together",
      "sub_themes": ["sub-theme name 1", "sub-theme name 2"]
    }}
  ]
}}"""


def run_turn1(group, obj, cat, clusters):
    """Turn 1: Propose 2-3 themes for a single objective×category slice."""
    group_label = GROUP_LABELS[group]
    obj_label = OBJ_LABELS[obj]
    cat_label = CAT_LABELS[cat]

    lines = []
    for c in clusters:
        lines.append(f'- [{c["coverage"]}] {c["name"]}')
        if c['findings']:
            # Show first few findings as context
            findings_short = c['findings'][:200]
            lines.append(f'    Findings: {findings_short}')

    user = f"""PARTICIPANT GROUP: {group_label}
OBJECTIVE: {obj_label}
CATEGORY: {cat_label}
NUMBER OF SUB-THEMES: {len(clusters)}

SUB-THEMES (sorted by coverage):
{chr(10).join(lines)}

Based ONLY on these sub-themes, propose 2-3 broad themes. Each theme must group at least 2 sub-themes. Every sub-theme must be assigned to exactly one theme. Stay grounded — do not add concepts absent from this data."""

    label = f'{group}/{obj}/{cat}'
    print(f'    Turn 1 [{label}]: {len(clusters)} sub-themes → ', end='', flush=True)
    data, t_in, t_out = llm_call(TURN1_SYSTEM, user)
    themes = data.get('themes', [])
    print(f'{len(themes)} themes ({t_in}+{t_out} tokens)')
    for t in themes:
        print(f'      • {t["name"]}')
    return themes


# ────────────────────────────────────────
# Turn 2: Consolidate across objectives
# ────────────────────────────────────────

TURN2_SYSTEM = f"""{STUDY_CONTEXT}

TASK: You have received theme proposals from 3 separate objectives (Early Detection, Diagnosis & Treatment, Continuity & Follow-Up) for the same participant group and category. Some proposed themes may overlap or be very similar across objectives.

Your job: CONSOLIDATE these into 4-6 FINAL overarching themes for the entire section.

STRICT RULES:
1. MERGE OVERLAPS: If "Awareness and education" appears in Objective 1 and "Patient education and counselling" in Objective 3, merge into one theme.
2. PRESERVE DISTINCTNESS: Do not over-merge. If two themes cover genuinely different concepts, keep them separate.
3. GROUNDED: The final themes must still be traceable to the original sub-themes. Do not invent new concepts during consolidation.
4. TARGET 4-6 final themes. This is for a thesis — too many themes fragments the narrative, too few loses nuance.
5. NAMING: Final theme names should work as thesis section headings (5-15 words, academic but clear).
6. For each final theme, list ALL the sub-themes (from all 3 objectives) that belong to it.
7. COUNT CHECK: The total number of sub-themes across all your final themes MUST equal exactly the number provided in the complete list. Count before responding.

OUTPUT FORMAT — return ONLY this JSON:
{{
  "final_themes": [
    {{
      "name": "Final consolidated theme name",
      "description": "2-3 sentence description of what this theme captures, suitable for a thesis",
      "sub_themes_included": ["sub-theme 1", "sub-theme 2", "sub-theme 3"]
    }}
  ]
}}"""


def run_turn2(group, cat, per_obj_themes, all_sub_theme_names):
    """Turn 2: Consolidate per-objective themes into 4-6 final themes."""
    group_label = GROUP_LABELS[group]
    cat_label = CAT_LABELS[cat]

    obj_sections = []
    for obj in ['objective_1', 'objective_2', 'objective_3']:
        themes = per_obj_themes.get(obj, [])
        if not themes:
            continue
        obj_sections.append(f'\n--- {OBJ_LABELS[obj]} ---')
        for t in themes:
            subs = ', '.join(t.get('sub_themes', []))
            obj_sections.append(f'  Theme: "{t["name"]}"')
            obj_sections.append(f'    Sub-themes: {subs}')
            obj_sections.append(f'    Rationale: {t.get("rationale", "")}')

    all_subs_list = '\n'.join(f'  - {s}' for s in sorted(all_sub_theme_names))

    user = f"""PARTICIPANT GROUP: {group_label}
CATEGORY: {cat_label}

PROPOSED THEMES BY OBJECTIVE:
{chr(10).join(obj_sections)}

COMPLETE LIST OF ALL SUB-THEMES TO COVER ({len(all_sub_theme_names)} total):
{all_subs_list}

INSTRUCTIONS:
1. Review the proposed themes across all 3 objectives.
2. Merge overlapping themes into unified final themes.
3. Ensure EVERY sub-theme from the complete list above is assigned to exactly one final theme.
4. Produce 4-6 final themes.
5. If a sub-theme was not covered by any proposed theme, assign it to the most fitting final theme anyway — do not drop it."""

    label = f'{group}/{cat}'
    print(f'    Turn 2 [{label}]: Consolidating → ', end='', flush=True)
    data, t_in, t_out = llm_call(TURN2_SYSTEM, user, max_tokens=6144)
    final = data.get('final_themes', [])
    total_assigned = sum(len(t.get('sub_themes_included', [])) for t in final)
    print(f'{len(final)} final themes, {total_assigned}/{len(all_sub_theme_names)} sub-themes assigned ({t_in}+{t_out} tokens)')
    for t in final:
        print(f'      • {t["name"]} ({len(t.get("sub_themes_included", []))} sub-themes)')
    return final


# ────────────────────────────────────────
# Turn 3: Assign clusters + pick excerpts
# ────────────────────────────────────────

TURN3_SYSTEM = f"""{STUDY_CONTEXT}

TASK: You have a finalised set of broad themes and the full list of sub-themes (clusters) with their findings text. Your job is to produce the final assignment: map each sub-theme to its theme, and select 1-2 representative participant quotes from the findings.

STRICT RULES:
1. EVERY sub-theme must appear in exactly one theme. Count carefully — do not drop any.
2. QUOTES ONLY FROM DATA: Select excerpts from the "Findings" text provided for each sub-theme. Copy them exactly. Do NOT rephrase, paraphrase, or invent quotes.
3. If a sub-theme's findings text seems mismatched, unclear, or irrelevant to the sub-theme name, use the sub-theme name itself as the quote instead of copying incorrect data.
4. Keep quotes concise — select the most impactful 1-2 quotes per sub-theme (max 150 characters each).
5. You are processing ONE OBJECTIVE at a time. Only assign sub-themes from this objective's data.

OUTPUT FORMAT — return ONLY this JSON:
{{
  "assignments": [
    {{
      "theme": "Exact final theme name",
      "sub_themes": [
        {{
          "name": "Exact sub-theme/cluster name",
          "selected_quotes": ["exact quote from findings data", "another quote if available"],
          "coverage": "e.g. 5/16",
          "participants": "e.g. P-001; P-005; S-002"
        }}
      ]
    }}
  ]
}}"""


def run_turn3_for_objective(group, cat, obj, final_themes, clusters):
    """Turn 3: Map clusters from ONE objective to a final theme + extract quotes."""
    group_label = GROUP_LABELS[group]
    cat_label = CAT_LABELS[cat]
    obj_label = OBJ_LABELS[obj]

    themes_list = '\n'.join(f'  {i+1}. {t["name"]}' for i, t in enumerate(final_themes))

    cluster_details = []
    for c in clusters:
        findings_trunc = c['findings'][:300] if c['findings'] else '(no findings)'
        cluster_details.append(
            f'SUB-THEME: {c["name"]}\n'
            f'  Coverage: {c["coverage"]}\n'
            f'  Participants: {c["participants"]}\n'
            f'  Findings: {findings_trunc}'
        )

    user = f"""PARTICIPANT GROUP: {group_label}
CATEGORY: {cat_label}
OBJECTIVE: {obj_label}

FINAL THEMES (use these for assignment — they span all objectives):
{themes_list}

SUB-THEMES TO ASSIGN FROM THIS OBJECTIVE ({len(clusters)} total):

{chr(10).join(cluster_details)}

INSTRUCTIONS:
1. Assign each sub-theme to exactly one of the final themes listed above.
2. For each sub-theme, select 1-2 short quotes from its Findings text. Copy exactly — do not rephrase.
3. If the Findings text seems mismatched with the sub-theme name (e.g. findings about finances under a clinical sub-theme), use the sub-theme name as the quote instead.
4. Every sub-theme must appear in your output. COUNT CHECK: your output must contain exactly {len(clusters)} sub-themes total.
5. Preserve the coverage and participants data for each sub-theme."""

    label = f'{group}/{obj}/{cat}'
    print(f'    Turn 3 [{label}]: Assigning {len(clusters)} sub-themes → ', end='', flush=True)
    data, t_in, t_out = llm_call(TURN3_SYSTEM, user, max_tokens=8192)
    assignments = data.get('assignments', [])
    total = sum(len(a.get('sub_themes', [])) for a in assignments)
    print(f'{total}/{len(clusters)} assigned ({t_in}+{t_out} tokens)')

    if total < len(clusters):
        print(f'      ⚠ WARNING: {len(clusters) - total} sub-themes dropped!')

    return assignments


def merge_assignments(all_assignment_batches):
    """Merge Turn 3 results from multiple objectives into a single assignment list."""
    theme_map = {}  # theme_name -> { theme, sub_themes: [] }
    for batch in all_assignment_batches:
        for entry in batch:
            tname = entry.get('theme', '')
            if tname not in theme_map:
                theme_map[tname] = {'theme': tname, 'sub_themes': []}
            theme_map[tname]['sub_themes'].extend(entry.get('sub_themes', []))
    return list(theme_map.values())


# ══════════════════════════════════════════════════════════════
# STEP 3: Process all 4 sections (with checkpoint/resume)
# ══════════════════════════════════════════════════════════════

SECTIONS = [
    ('patient_survivor', 'facilitator', 'P+S Facilitators', '2E7D32'),
    ('patient_survivor', 'barrier', 'P+S Barriers', 'C62828'),
    ('doctor', 'facilitator', 'Doc Facilitators', '1565C0'),
    ('doctor', 'barrier', 'Doc Barriers', 'E65100'),
]

CHECKPOINT_PATH = ROOT / 'themed-analysis-checkpoint.json'


def load_checkpoint():
    if CHECKPOINT_PATH.exists():
        print(f'📂 Resuming from checkpoint: {CHECKPOINT_PATH}')
        return json.loads(CHECKPOINT_PATH.read_text())
    return {}


def save_checkpoint(data):
    CHECKPOINT_PATH.write_text(json.dumps(data, indent=2, ensure_ascii=False, default=str))


checkpoint = load_checkpoint()
results = {}
total_calls = 0
skipped_calls = 0

for group, cat, sheet_name, tab_color in SECTIONS:
    section_key = f'{group}_{cat}'
    print(f'\n{"="*60}')
    print(f'Section: {GROUP_LABELS[group]} — {CAT_LABELS[cat]}')
    print(f'{"="*60}')

    # Check if this entire section is already done
    if section_key in checkpoint and 'assignments' in checkpoint[section_key]:
        print(f'  ✅ Already completed (loaded from checkpoint)')
        results[section_key] = checkpoint[section_key]
        skipped_calls += 1
        continue

    # Collect all clusters for this section across objectives
    all_clusters = []
    per_obj_themes = checkpoint.get(section_key, {}).get('per_obj_themes', {})

    for obj in ['objective_1', 'objective_2', 'objective_3']:
        clusters = parsed_sections.get((group, obj, cat), [])
        if not clusters:
            print(f'  {OBJ_LABELS[obj]}: no data, skipping')
            continue

        print(f'  {OBJ_LABELS[obj]}: {len(clusters)} sub-themes')
        all_clusters.extend(clusters)

        # Turn 1: propose themes for this slice (skip if already in checkpoint)
        if obj in per_obj_themes:
            print(f'    Turn 1 [{group}/{obj}/{cat}]: ✅ cached from checkpoint')
            skipped_calls += 1
        else:
            themes = run_turn1(group, obj, cat, clusters)
            per_obj_themes[obj] = themes
            total_calls += 1
            # Save checkpoint after each Turn 1
            checkpoint.setdefault(section_key, {})['per_obj_themes'] = per_obj_themes
            checkpoint[section_key]['sheet_name'] = sheet_name
            checkpoint[section_key]['tab_color'] = tab_color
            save_checkpoint(checkpoint)

    if not all_clusters:
        print(f'  ⚠ No data for this section, skipping')
        continue

    all_sub_theme_names = [c['name'] for c in all_clusters]

    # Turn 2: consolidate (skip if already in checkpoint)
    cached_final = checkpoint.get(section_key, {}).get('final_themes')
    if cached_final:
        print(f'    Turn 2 [{group}/{cat}]: ✅ cached from checkpoint')
        final_themes = cached_final
        skipped_calls += 1
    else:
        final_themes = run_turn2(group, cat, per_obj_themes, all_sub_theme_names)
        total_calls += 1
        checkpoint.setdefault(section_key, {})['final_themes'] = final_themes
        save_checkpoint(checkpoint)

    # Turn 3: assign + excerpts (split per objective, skip already-done objectives)
    cached_t3 = checkpoint.get(section_key, {}).get('turn3_batches', {})
    assignment_batches = []
    for obj in ['objective_1', 'objective_2', 'objective_3']:
        obj_clusters = parsed_sections.get((group, obj, cat), [])
        if not obj_clusters:
            continue
        if obj in cached_t3:
            print(f'    Turn 3 [{group}/{obj}/{cat}]: ✅ cached from checkpoint')
            assignment_batches.append(cached_t3[obj])
            skipped_calls += 1
        else:
            batch = run_turn3_for_objective(group, cat, obj, final_themes, obj_clusters)
            assignment_batches.append(batch)
            total_calls += 1
            # Save after each Turn 3 sub-call
            checkpoint.setdefault(section_key, {}).setdefault('turn3_batches', {})[obj] = batch
            save_checkpoint(checkpoint)

    assignments = merge_assignments(assignment_batches)

    results[section_key] = {
        'per_obj_themes': {k: v for k, v in per_obj_themes.items()},
        'final_themes': final_themes,
        'assignments': assignments,
        'sheet_name': sheet_name,
        'tab_color': tab_color,
    }
    # Mark section as fully done in checkpoint
    checkpoint[section_key] = results[section_key]
    save_checkpoint(checkpoint)

print(f'\n\nTotal LLM calls: {total_calls} (skipped {skipped_calls} from checkpoint)')

# ══════════════════════════════════════════════════════════════
# STEP 4: Save raw JSON
# ══════════════════════════════════════════════════════════════

json_path = ROOT / 'themed-analysis-raw.json'
json_path.write_text(json.dumps(results, indent=2, ensure_ascii=False, default=str))
print(f'Saved raw LLM output to: {json_path}')

# ══════════════════════════════════════════════════════════════
# STEP 5: Build output Excel
# ══════════════════════════════════════════════════════════════

# Styles
theme_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
theme_font = Font(name='Calibri', bold=True, size=11, color='1A1F2C')
code_font = Font(name='Calibri', size=10, color='4A5263')
excerpt_font = Font(name='Calibri', size=9, italic=True, color='6B7280')
header_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
wrap = Alignment(wrap_text=True, vertical='top')
thin_border = Border(
    bottom=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
)


def write_sheet(wb, sheet_name, tab_color, assignments):
    """Write a themed analysis sheet."""
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = tab_color

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 48
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 28

    r = 1
    for col, lbl in enumerate(['Theme', 'Sub-themes / Codes', 'Participant Quotes', 'Coverage', 'Participants'], 1):
        c = ws.cell(row=r, column=col, value=lbl)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border
    ws.row_dimensions[r].height = 24
    r += 1

    for theme_group in assignments:
        theme_name = theme_group.get('theme', '')
        subs = theme_group.get('sub_themes', [])
        if not subs:
            continue

        theme_start = r

        for si, sub in enumerate(subs):
            sub_name = sub.get('name', '')
            quotes = sub.get('selected_quotes', [])
            coverage = sub.get('coverage', '')
            participants = sub.get('participants', '')

            quotes_text = '\n'.join(f'"{q}"' for q in quotes[:2]) if quotes else '—'

            if si == 0:
                c = ws.cell(row=r, column=1, value=theme_name)
                c.font = theme_font
                c.fill = theme_fill
                c.alignment = wrap
                c.border = thin_border

            c = ws.cell(row=r, column=2, value=sub_name)
            c.font = code_font
            c.alignment = wrap
            c.border = thin_border

            c = ws.cell(row=r, column=3, value=quotes_text)
            c.font = excerpt_font
            c.alignment = wrap
            c.border = thin_border

            c = ws.cell(row=r, column=4, value=coverage)
            c.font = code_font
            c.alignment = Alignment(horizontal='center', vertical='top')
            c.border = thin_border

            c = ws.cell(row=r, column=5, value=participants)
            c.font = code_font
            c.alignment = wrap
            c.border = thin_border

            lines = max(len(quotes), 1)
            ws.row_dimensions[r].height = max(30, lines * 18)
            r += 1

        if len(subs) > 1:
            ws.merge_cells(start_row=theme_start, start_column=1,
                           end_row=theme_start + len(subs) - 1, end_column=1)
        r += 1

    ws.freeze_panes = 'A2'


wb_out = Workbook()
wb_out.remove(wb_out.active)

for group, cat, sheet_name, tab_color in SECTIONS:
    section_key = f'{group}_{cat}'
    if section_key not in results:
        continue
    write_sheet(wb_out, sheet_name, tab_color, results[section_key]['assignments'])

excel_path = ROOT / 'thematic-analysis.xlsx'
wb_out.save(str(excel_path))
print(f'\n✅ Written: {excel_path}')
print(f'   Sheets: {[s for _, _, s, _ in SECTIONS if f"{_}_{__}" in results] if False else list(results.keys())}')
